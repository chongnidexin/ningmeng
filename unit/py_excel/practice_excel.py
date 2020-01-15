from openpyxl import load_workbook

# 存到Excel 里面   python去操作Excel
# 1.只支持这种后缀 xlsx--->openpyxl
# 2.老老实实的创建


# load_workbook 使用方式
# 1.打开Excel
wb = load_workbook("test.xlsx")
# 2.定位表单
sheet = wb['test']  # 传表单名 返回一个表单对象
print('最大行：{}'.format(sheet.max_row))   #表单的最大行
print('最大列：{}'.format(sheet.max_column)) # 表单的最大列

# 定位单元格 行和列
res = sheet.cell(1, 1).value
# 数据从excel 里拿出来的类型  数字还是数字，其余都是字符串
print("拿到的结果是：{}".format(res))
