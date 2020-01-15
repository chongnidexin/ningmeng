from openpyxl import load_workbook


class DoExcel:
    """定义获取文件数据类"""

    def __init__(self, file_name, sheet_name):
        """"定义初始化函数，传递文件名及工作簿名"""
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        """定义获取文件数据函数"""

        wb = load_workbook(self.file_name)  # 打开excel
        sheet = wb[self.sheet_name]  # 定位工作簿 sheet
        test_data = []
        for i in range(1, sheet.max_row + 1):

            sub_data = {'method': sheet.cell(i, 1).value, 'url': sheet.cell(i, 2).value, 'data': sheet.cell(i, 3).value,
                        'expected': sheet.cell(i, 4).value}
            # sub_data = {}
            # # method = sheet.cell(1, 1).value  # 定义变量接收 1,1 单元格的值
            # sub_data['method'] = sheet.cell(i, 1).value
            # # url = sheet.cell(1,2).value
            # sub_data['url'] = sheet.cell(i, 2).value
            # # data = sheet.cell(1,3).value
            # sub_data['data'] = sheet.cell(i, 3).value
            # # expected = sheet.cell(1,4).value
            # sub_data['expected'] = sheet.cell(i, 4).value

            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    print(DoExcel('test.xlsx', 'test').get_data())
